import time
import os
import argparse
import xlsxwriter
import concurrent.futures
import openpyxl
import requests
from datetime import datetime
from selenium import webdriver
from pathlib import Path
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

BASE_DIR = Path(__file__).parent.parent
MEDIA_DIR = BASE_DIR / "media"
os.makedirs(MEDIA_DIR, exist_ok=True)

def send_telegram(text: str):
    token = "token"
    url = "https://api.telegram.org/bot"
    channel_id = "@scrgmv1"
    url += token
    method = url + "/sendMessage"

    r = requests.post(method, data={
        "chat_id": channel_id,
        "text": text
    })

    if r.status_code != 200:
        raise Exception("post_text error")


if not os.path.exists('output'):
    os.makedirs('output')
name_folder = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
os.mkdir('output/' + name_folder)

parser = argparse.ArgumentParser(description='Scrape businesses from Google Maps')
parser.add_argument('--business_type', '-bt', '-business_type', '--bt')

args = parser.parse_args()

business_type = args.business_type

locations = [
    # 'Украина Киев',
    # 'Украина Харьков',
    # 'Украина Днепр',
    # 'Украина Одесса',
    # 'Украина Запорожье',
    # 'Украина Львов',
    # 'Украина Кривой Рог',
    # 'Украина Николаев',
    # 'Украина Винница',
    # 'Украина Херсон',
    # 'Украина Полтава',
    # 'Украина Чернигов',
    # 'Украина Черкассы',
    # 'Украина Хмельницкий',
    # 'Украина Житомир',
    # 'Украина Сумы',
    # 'Украина Ровно',
    # 'Украина Ивано-Франковск',
    # 'Украина Тернополь',
    # 'Украина Каменец-Подольский',
    # 'Украина Луцк',
    # 'Украина Краматорск',
    # 'Украина Ужгород',
    # 'Украина Бровары',
    # 'Украина Дружковка',
    # 'Украина Павлоград',
    # 'Украина Черновцы',
    # 'Украина Буча',
    # 'Украина Конотоп',
    # 'Украина Никополь',
    # 'Украина Борисполь',
    # 'Украина Кропивницкий',
    # 'Украина Вознесенск',
    # 'Украина Дрогобыч',
    # 'Украина Желтые Воды',
    # 'Украина Знаменка',
    # 'Украина Измаил',
    # 'Украина Калуш',
    # 'Украина Каменское',
    # 'Украина Кобеляки',
    # 'Украина Коломыя',
    # 'Украина Кременчуг',
    # 'Украина Лозовая',
    # 'Украина Лубны',
    # 'Украина Новоград-Волынский',
    # 'Украина Пятихатки',
    # 'Украина Славутич',
    # 'Украина Фастов',
    # 'Украина Черноморск',
    # 'Украина Чоп',
    # 'Украина Шостка',
    # 'Украина Шпола',
    # 'Украина Новомосковск',
    'Украина Южноукраинск',
    # 'Украина Самбор',
    # 'Украина Жовква',
    # 'Украина Винники',
    # 'Украина Мерефа',
    # 'Украина Чугуев',
    # 'Украина Ирпень',
    # 'Украина Белая Церковь',
    # 'Украина Коростень',
    # 'Украина Новоград-Волинський',
    # 'Украина Новая Одесса',
    # 'Украина Трускавец',
    # 'Украина Яготин',
    # 'Украина Славянск',
    # 'Украина Змиев',
    # 'Украина Новая Водолага',
    # 'Украина Миргород',
    # 'Украина Ахтырка',
    # 'Украина Бахмач',
    # 'Украина Южный',
    # 'Украина Килия',
    # 'Украина Болград',
    # 'Украина Арциз',
    # 'Украина Подол',
    # 'Украина Подольск',
    # 'Украина Балта',
    # 'Украина Збараж',
    # 'Украина Дубно',
    # 'Украина Городок',
    # 'Украина Яворов',
    # 'Украина Владимир',
    # 'Украина Нововолынск',
    # 'Украина Киверцы',
    # 'Украина Ковель',
    # 'Украина Рогатин',
    # 'Украина Бурштин',
    # 'Украина Галич',
    # 'Украина Переяслав',
    # 'Украина Пирятин',
    # 'Украина Лохвиця',
    # 'Украина Кобеляки',
    # 'Украина Верхнеднепровск',
    # 'Украина Близнюки',
    # 'Украина Богодухов',
    # 'Украина Красноград',
    # 'Украина Валки',
    # 'Украина Карловка',
    # 'Украина Гадяч',
]


print(f"Searching for {business_type} in {locations}")


def scrape_location(location):
    print(f"Scraping {business_type} in {location}")

    driver = webdriver.Chrome()

    driver.get('https://www.google.com/maps')


    try:
        search_box = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'searchboxinput'))
        )

        search_query = f'{business_type} in {location}'
        search_box.send_keys(search_query)
        search_box.send_keys(Keys.RETURN)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'hfpxzc'))
        )
    except:
        time.sleep(5)
        try:
            search_box = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, 'searchboxinput'))
            )

            search_query = f'{business_type} in {location}'
            search_box.send_keys(search_query)
            search_box.send_keys(Keys.RETURN)

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'hfpxzc'))
            )
        except:
            print(f"Error: Failed to load search results for {location}")
            driver.quit()
            return

    output = []
    urls = []

    while True:
        try:
            businesses = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, 'hfpxzc'))
            )
        except:
            break

        time.sleep(2)

        for business in businesses:
            url = business.get_attribute('href')
            if url not in urls:
                urls.append(url)

        driver.execute_script("arguments[0].scrollIntoView();", businesses[-1])
        time.sleep(2)

        try:
            new_businesses = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, 'hfpxzc'))
            )
            if len(new_businesses) == len(businesses):
                break
        except:
            break

    for url in urls:
        driver.get(url)

        try:
            name = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'DUwDvf'))
            ).text
        except Exception as e:
            name = ""

        try:
            phone = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "[data-tooltip*='phone']"))
            ).text
        except Exception as e:
            phone = ""

        try:
            address = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "[data-item-id*='address']"))
            ).text
        except Exception as e:
            address = ""

        output.append([name, phone, address])

        for item in output:
            if not item[1]:
                output.remove(item)

    output_file = f'output/{name_folder}/{business_type}_{location}.xlsx'

    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet()

    header_format = workbook.add_format({'bold': True})
    header_row = ['Name', 'Phone', 'Address']
    for col_num, value in enumerate(header_row):
        worksheet.write(0, col_num, value, header_format)

    for row_num, row_data in enumerate(output, start=1):
        for col_num, value in enumerate(row_data):
            worksheet.write(row_num, col_num, value)

    workbook.close()

    driver.quit()


threads = []
interval = 4

with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
    for i in range(0, len(locations), interval):
        subarray = locations[i:i + interval]
        for location in subarray:
            thread = executor.submit(scrape_location, location)
            threads.append(thread)

output_file_merge = MEDIA_DIR / f'{name_folder}_merge.xlsx'

workbook = xlsxwriter.Workbook(output_file_merge)
worksheet = workbook.add_worksheet()
row_counter = 0

folder_path = 'output/' + name_folder
if not os.path.exists(folder_path):
    print(f"Directory {folder_path} does not exist.")
else:
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(folder_path, file_name)
            wb = openpyxl.load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    row_values = [cell.value for cell in row]
                    worksheet.write_row(row_counter, 0, row_values)
                    row_counter += 1

workbook.close()

for file_name in os.listdir("output/" + name_folder):
    file_path = os.path.join("output/" + name_folder, file_name)
    os.remove(file_path)

os.rmdir("output/" + name_folder)

#send_telegram("Парсер закончил работу над: " + business_type + '. Файл называется: ' + name_folder + '_merge.xlsx')
